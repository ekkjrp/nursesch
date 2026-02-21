"""근무표 라우터 — 생성/조회/수정/확정/통계/Excel 내보내기"""
import json
import io
from typing import List, Optional
from datetime import date as dt_date
import calendar
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app import models, schemas
from app.auth import get_current_nurse, require_admin
from app.engine.scheduler import generate_schedule
from app.engine.validator import validate_schedule, _is_weekend

router = APIRouter(prefix="/api/schedules", tags=["schedules"])

SHIFT_KR = {"D": "데이", "E": "이브닝", "N": "나이트", "M": "미드", "O": "오프", "Y": "연차", "X": "기타"}


def _get_rule_dict(rule: models.Rule) -> dict:
    return {
        "max_consecutive_work_days": rule.max_consecutive_work_days,
        "max_consecutive_night_shifts": rule.max_consecutive_night_shifts,
        "min_monthly_off_days": rule.min_monthly_off_days,
        "weekday_day_charge": rule.weekday_day_charge,
        "weekday_day_action": rule.weekday_day_action,
        "weekday_evening_charge": rule.weekday_evening_charge,
        "weekday_evening_action": rule.weekday_evening_action,
        "weekday_night_charge": rule.weekday_night_charge,
        "weekday_night_action": rule.weekday_night_action,
        "weekend_day_charge": rule.weekend_day_charge,
        "weekend_day_action": rule.weekend_day_action,
        "weekend_evening_charge": rule.weekend_evening_charge,
        "weekend_evening_action": rule.weekend_evening_action,
        "weekend_night_charge": rule.weekend_night_charge,
        "weekend_night_action": rule.weekend_night_action,
        "allow_night_to_day": rule.allow_night_to_day,
        "allow_night_to_evening": rule.allow_night_to_evening,
        "allow_night_off_day": rule.allow_night_off_day,
        "an_auto_weekend_off": rule.an_auto_weekend_off,
        "max_monthly_night_shifts": rule.max_monthly_night_shifts or 0,
    }


@router.get("", response_model=List[schemas.ScheduleOut])
def list_schedules(
    ward_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_nurse),
):
    q = db.query(models.Schedule)
    if ward_id:
        q = q.filter(models.Schedule.ward_id == ward_id)
    return q.order_by(models.Schedule.year_month.desc()).all()


@router.post("/generate")
def generate(
    body: schemas.ScheduleGenerateRequest,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    ward = db.query(models.Ward).filter(models.Ward.id == body.ward_id).first()
    if not ward:
        raise HTTPException(status_code=404, detail="병동을 찾을 수 없습니다.")

    rule = db.query(models.Rule).filter(models.Rule.ward_id == body.ward_id).first()
    if not rule:
        raise HTTPException(status_code=400, detail="근무 규칙이 설정되지 않았습니다.")

    nurses = db.query(models.Nurse).filter(models.Nurse.ward_id == body.ward_id).all()
    if not nurses:
        raise HTTPException(status_code=400, detail="병동에 간호사가 없습니다.")

    # 공휴일 목록
    holidays_db = db.query(models.Holiday).filter(models.Holiday.ward_id == body.ward_id).all()
    holidays = [h.date for h in holidays_db]

    # 승인된 근무 요청
    approved_requests = db.query(models.ShiftRequest).filter(
        models.ShiftRequest.status == "APPROVED",
        models.ShiftRequest.date.startswith(body.year_month),
    ).all()
    req_dicts = [
        {"nurse_id": r.nurse_id, "date": r.date, "requested_shift_type": r.requested_shift_type}
        for r in approved_requests
    ]

    # 연차/희망 휴일 맵
    leaves = db.query(models.NurseMonthlyLeave).filter(
        models.NurseMonthlyLeave.year_month == body.year_month
    ).all()
    leave_map = {}
    for lv in leaves:
        leave_map[lv.nurse_id] = {
            "annual_leave_count": lv.annual_leave_count,
            "requested_off_dates": json.loads(lv.requested_off_dates or "[]"),
        }

    nurse_dicts = [
        {
            "id": n.id,
            "name": n.name,
            "grade": n.grade,
            "is_night_dedicated": n.is_night_dedicated,
            "dedicated_shift": n.dedicated_shift,
            "weekday_preference": n.weekday_preference,
            "weekend_preference": n.weekend_preference,
            "monthly_annual_leave": n.monthly_annual_leave,
            "sort_order": n.sort_order,
        }
        for n in nurses
    ]

    rule_dict = _get_rule_dict(rule)

    # 기존 DRAFT 삭제 후 재생성
    existing = db.query(models.Schedule).filter(
        models.Schedule.ward_id == body.ward_id,
        models.Schedule.year_month == body.year_month,
        models.Schedule.status == "DRAFT",
    ).first()
    if existing:
        db.delete(existing)
        db.commit()

    result = generate_schedule(
        nurses=nurse_dicts,
        rule=rule_dict,
        year_month=body.year_month,
        holidays=holidays,
        approved_requests=req_dicts,
        leave_map=leave_map,
        prev_entries=body.prev_month_entries,
    )

    schedule = models.Schedule(ward_id=body.ward_id, year_month=body.year_month, status="DRAFT")
    db.add(schedule)
    db.commit()
    db.refresh(schedule)

    for e in result["entries"]:
        entry = models.ScheduleEntry(
            schedule_id=schedule.id,
            nurse_id=e["nurse_id"],
            date=e["date"],
            shift_type=e["shift_type"],
        )
        db.add(entry)
    db.commit()

    return {
        "schedule_id": schedule.id,
        "score": result["score"],
        "hard_violations": sum(1 for v in result["violations"] if v.get("type") == "HARD"),
        "violations": result["violations"][:20],  # 최대 20건만 반환
    }


@router.get("/{schedule_id}")
def get_schedule(schedule_id: int, db: Session = Depends(get_db), _=Depends(get_current_nurse)):
    schedule = db.query(models.Schedule).filter(models.Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="근무표를 찾을 수 없습니다.")
    entries = db.query(models.ScheduleEntry).filter(
        models.ScheduleEntry.schedule_id == schedule_id
    ).all()
    return {
        "schedule": schemas.ScheduleOut.model_validate(schedule),
        "entries": [schemas.EntryOut.model_validate(e) for e in entries],
    }


@router.put("/{schedule_id}/entries/{entry_id}", response_model=schemas.EntryOut)
def update_entry(
    schedule_id: int,
    entry_id: int,
    body: schemas.EntryUpdate,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    entry = db.query(models.ScheduleEntry).filter(
        models.ScheduleEntry.id == entry_id,
        models.ScheduleEntry.schedule_id == schedule_id,
    ).first()
    if not entry:
        raise HTTPException(status_code=404, detail="항목을 찾을 수 없습니다.")
    entry.shift_type = body.shift_type
    entry.is_manually_edited = True
    db.commit()
    db.refresh(entry)
    return entry


@router.post("/{schedule_id}/confirm", response_model=schemas.ScheduleOut)
def confirm_schedule(schedule_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    schedule = db.query(models.Schedule).filter(models.Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="근무표를 찾을 수 없습니다.")
    schedule.status = "CONFIRMED"
    db.commit()
    db.refresh(schedule)
    return schedule


@router.get("/{schedule_id}/validate")
def validate(schedule_id: int, db: Session = Depends(get_db), _=Depends(get_current_nurse)):
    schedule = db.query(models.Schedule).filter(models.Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="근무표를 찾을 수 없습니다.")
    entries = db.query(models.ScheduleEntry).filter(
        models.ScheduleEntry.schedule_id == schedule_id
    ).all()
    rule = db.query(models.Rule).filter(models.Rule.ward_id == schedule.ward_id).first()
    nurses = db.query(models.Nurse).filter(models.Nurse.ward_id == schedule.ward_id).all()
    holidays_db = db.query(models.Holiday).filter(models.Holiday.ward_id == schedule.ward_id).all()

    entry_dicts = [{"nurse_id": e.nurse_id, "date": e.date, "shift_type": e.shift_type} for e in entries]
    nurse_dicts = [{"id": n.id, "grade": n.grade, "is_night_dedicated": n.is_night_dedicated, "dedicated_shift": n.dedicated_shift} for n in nurses]
    rule_dict = _get_rule_dict(rule) if rule else {}
    holiday_list = [h.date for h in holidays_db]

    return validate_schedule(entry_dicts, nurse_dicts, rule_dict, schedule.year_month, holiday_list)


@router.get("/{schedule_id}/stats", response_model=List[schemas.NurseStats])
def get_stats(schedule_id: int, db: Session = Depends(get_db), _=Depends(get_current_nurse)):
    schedule = db.query(models.Schedule).filter(models.Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="근무표를 찾을 수 없습니다.")
    nurses = db.query(models.Nurse).filter(models.Nurse.ward_id == schedule.ward_id).all()
    entries = db.query(models.ScheduleEntry).filter(
        models.ScheduleEntry.schedule_id == schedule_id
    ).all()

    entry_map: dict = {}
    for e in entries:
        entry_map.setdefault(e.nurse_id, []).append(e)

    result = []
    for nurse in nurses:
        ents = entry_map.get(nurse.id, [])
        stats = schemas.NurseStats(nurse_id=nurse.id, nurse_name=nurse.name, grade=nurse.grade)
        for e in ents:
            if e.shift_type == "D": stats.d_count += 1
            elif e.shift_type == "E": stats.e_count += 1
            elif e.shift_type == "N": stats.n_count += 1
            elif e.shift_type == "M": stats.m_count += 1
            elif e.shift_type == "O": stats.o_count += 1
            elif e.shift_type == "Y": stats.y_count += 1
            if e.shift_type not in ("O", "Y", "X") and _is_weekend(e.date):
                stats.weekend_work += 1
        stats.total_off = stats.o_count + stats.y_count
        result.append(stats)
    return result


@router.get("/{schedule_id}/export")
def export_excel(schedule_id: int, db: Session = Depends(get_db), _=Depends(get_current_nurse)):
    """Excel 내보내기"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl이 설치되지 않았습니다.")

    schedule = db.query(models.Schedule).filter(models.Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="근무표를 찾을 수 없습니다.")

    nurses = db.query(models.Nurse).filter(models.Nurse.ward_id == schedule.ward_id).order_by(
        models.Nurse.sort_order, models.Nurse.name
    ).all()
    entries = db.query(models.ScheduleEntry).filter(
        models.ScheduleEntry.schedule_id == schedule_id
    ).all()

    entry_map: dict = {}
    for e in entries:
        entry_map.setdefault(e.nurse_id, {})[e.date] = e.shift_type

    year, month = map(int, schedule.year_month.split("-"))
    days_in_month = calendar.monthrange(year, month)[1]
    all_dates = [f"{schedule.year_month}-{d:02d}" for d in range(1, days_in_month + 1)]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{schedule.year_month} 근무표"

    COLORS = {"D": "4A90D9", "E": "7BC67E", "N": "9B59B6", "M": "5DADE2",
              "O": "BDC3C7", "Y": "E67E22", "X": "F1C40F"}

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 헤더
    ws.cell(1, 1, "이름").font = Font(bold=True)
    ws.cell(1, 2, "직급").font = Font(bold=True)
    for i, date in enumerate(all_dates):
        d = dt_date.fromisoformat(date)
        ws.cell(1, i + 3, f"{d.day}\n{['월','화','수','목','금','토','일'][d.weekday()]}")
        ws.cell(1, i + 3).alignment = Alignment(wrap_text=True, horizontal="center")
    # 통계 헤더
    for i, label in enumerate(["D", "E", "N", "M", "O", "Y", "합계"]):
        ws.cell(1, len(all_dates) + 3 + i, label).font = Font(bold=True)

    # 데이터
    for row, nurse in enumerate(nurses, start=2):
        ws.cell(row, 1, nurse.name)
        ws.cell(row, 2, nurse.grade)
        counts = {k: 0 for k in ["D", "E", "N", "M", "O", "Y"]}
        for col, date in enumerate(all_dates):
            st = entry_map.get(nurse.id, {}).get(date, "O")
            cell = ws.cell(row, col + 3, st)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
            if st in COLORS:
                cell.fill = PatternFill("solid", fgColor=COLORS[st])
            if st in counts:
                counts[st] += 1
        for i, key in enumerate(["D", "E", "N", "M", "O", "Y"]):
            ws.cell(row, len(all_dates) + 3 + i, counts[key])
        ws.cell(row, len(all_dates) + 9, sum(counts.values()))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"schedule_{schedule.year_month}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
